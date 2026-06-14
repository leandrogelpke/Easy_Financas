#!/usr/bin/env python3
"""
fetch-totvs.py — Parser cumulativo de relatórios de Comissão Totvs (ASE806).

Lê arquivos .eml ou .xlsx de um diretório de entrada e gera um snapshot JSON
cumulativo, deduplicando por hash do conteúdo. Estrutura cada documento como:
    tipo: BASE | COMPLEMENTAR | PROVISAO
    competencia: YYYY-MM-01
    cnpjs_filial: [{cnpj, razao_social, total}]
    clientes: [{cliente, cnpj_filial, valor_comissao, valor_bruto_item, produto, contrato, ...}]
    total_geral
    source_file, source_email_date, hash

Uso:
    python3 fetch-totvs.py --in <pasta_input> --out <snapshot.json> [--quiet]

Diretórios default:
    --in:   ~/Documents/Easy_Financas/totvs/raw  (onde caem os arquivos do Drive)
    --out:  ~/Documents/GRAP/.../bling-api/totvs_snapshot.json

Tipos de arquivo aceitos:
    - .eml: extrai .xlsx anexo
    - .xlsx: lê direto
"""
from __future__ import annotations

import argparse
import email
import glob
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from email import policy
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("erro: openpyxl ausente. instale com `pip3 install openpyxl --break-system-packages`", file=sys.stderr)
    sys.exit(1)

HOME = Path.home()
DEFAULT_IN = HOME / "Documents" / "Easy_Financas" / "totvs" / "raw"
DEFAULT_OUT = HOME / "Documents" / "GRAP" / "Negociação Easy" / "Controles Easy" / "relatorios atuais" / "bling-api" / "totvs_snapshot.json"


def log(msg, quiet=False):
    if not quiet:
        print(msg, file=sys.stderr)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def detect_tipo(subject: str, filename: str) -> str:
    txt = f"{subject or ''} {filename or ''}".lower()
    if "provis" in txt:
        return "PROVISAO"
    if "complement" in txt:
        return "COMPLEMENTAR"
    return "BASE"


def parse_competencia_from_filename(fn: str) -> Optional[str]:
    # ex: "ASE806 - EASY ... - 01.04.2026.xlsx" -> 2026-04-01
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", fn)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    # ex: "PROVISAO 04.2026.xlsx" -> 2026-04-01
    m = re.search(r"PROVISAO\s+(\d{2})\.(\d{4})", fn, re.IGNORECASE)
    if m:
        return f"{m.group(2)}-{m.group(1)}-01"
    return None


def extract_xlsx_from_eml(eml_path: Path, out_dir: Path) -> List[Tuple[Path, Dict[str, str]]]:
    """Extrai os anexos .xlsx do .eml para out_dir; retorna [(xlsx_path, meta)].
    Usa stem do .eml como prefixo para evitar colisão entre dois .emls com
    anexo de mesmo nome. Idempotente: se o .xlsx já existe em out_dir com
    o nome calculado, não reescreve (evita inflar .extracted/ a cada rodada).
    Meta inclui 'original_xlsx_filename' (nome do anexo conforme veio no .eml)
    para permitir match contra o manifest do Drive."""
    out_dir.mkdir(parents=True, exist_ok=True)
    with open(eml_path, "rb") as fh:
        msg = email.message_from_binary_file(fh, policy=policy.default)
    base_meta = {
        "subject": msg.get("Subject", "") or "",
        "from": msg.get("From", "") or "",
        "date": msg.get("Date", "") or "",
    }
    eml_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", eml_path.stem)[:60]
    results: List[Tuple[Path, Dict[str, str]]] = []
    for idx, part in enumerate(msg.iter_attachments()):
        fn = part.get_filename()
        if fn and fn.lower().endswith(".xlsx"):
            # nome único = stem do .eml + índice + nome original
            safe_fn = f"{eml_stem}__{idx}__{fn}"
            outp = out_dir / safe_fn
            if not outp.exists():
                data = part.get_payload(decode=True)
                outp.write_bytes(data)
            meta = dict(base_meta, original_xlsx_filename=fn)
            results.append((outp, meta))
    return results


def br_email_date_to_iso(s: str) -> Optional[str]:
    if not s:
        return None
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(s)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None


def is_faturamento_file(filename: str) -> bool:
    """FATURAMENTO TOTVS tem estrutura diferente: aba 'BASE' única (sem Plan2),
    pivot caches que quebram openpyxl normal, e contém múltiplos meses agregados."""
    return "FATURAMENTO" in filename.upper()


def parse_faturamento_xlsx(xlsx_path: Path, source_meta: Dict[str, str]) -> List[Dict[str, Any]]:
    """Parse 'FATURAMENTO TOTVS *.xlsx' — relatório com todos os meses realizados.
    Retorna uma lista de docs (um por REFERENCIA / YYYY-MM), tipo='REAL'.
    Usa read_only=True para contornar bug do openpyxl com pivot caches."""
    from collections import defaultdict
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "BASE" not in wb.sheetnames:
        raise ValueError(f"esperava aba 'BASE' em {xlsx_path.name}")
    sheet = wb["BASE"]

    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError(f"aba BASE vazia em {xlsx_path.name}")
    headers = {h: i for i, h in enumerate(header_row) if h}

    def get(row, key):
        i = headers.get(key)
        return row[i] if i is not None and i < len(row) else None

    def to_date(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10] if v else None

    def to_float(v):
        if v is None:
            return None
        try:
            return round(float(v), 2)
        except Exception:
            return None

    # Agrupa linhas por REFERENCIA (YYYY-MM-01)
    by_ref_clientes: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_ref_filiais: Dict[str, Dict[str, Dict[str, Any]]] = defaultdict(dict)
    skipped_invalid_ref = 0
    for row in rows_iter:
        if get(row, "VLRCOMIS") is None and get(row, "CLIENTE") is None:
            continue
        ref_raw = to_date(get(row, "REFERENCIA"))
        if not ref_raw:
            continue
        # Filtra linhas onde REFERENCIA virou '########' (truncamento Excel)
        # ou outro formato não-parseável como data YYYY-MM-DD
        if not (len(ref_raw) == 10 and ref_raw[4] == '-' and ref_raw[7] == '-'
                and ref_raw[:4].isdigit() and ref_raw[5:7].isdigit()):
            skipped_invalid_ref += 1
            continue
        ref_key = ref_raw[:7] + "-01"  # YYYY-MM-01

        cnpj = str(get(row, "CNPJ") or "")
        cod_empresa = str(get(row, "EMPRESA") or "")
        razao = str(get(row, "NOMEAR") or get(row, "CLIENTE") or "").strip()
        vlr_comis = to_float(get(row, "VLRCOMIS")) or 0.0

        # Acumula totalização por filial dentro deste mês
        filial_key = cnpj
        if filial_key not in by_ref_filiais[ref_key]:
            by_ref_filiais[ref_key][filial_key] = {
                "cod_empresa": cod_empresa,
                "cnpj": cnpj,
                "total": 0.0,
                "razao_social": razao,
            }
        by_ref_filiais[ref_key][filial_key]["total"] += vlr_comis

        # Detalhe por cliente (mesma estrutura do parser de comissão para
        # garantir compatibilidade com o build-html.py)
        by_ref_clientes[ref_key].append({
            "cnpj_filial": cnpj,
            "cod_cliente": str(get(row, "CODCLIENTE") or ""),
            "cliente": str(get(row, "CLIENTE") or "").strip(),
            "produto_cod": str(get(row, "PRODUTO") or ""),
            "produto_desc": str(get(row, "DESCPROD") or "").strip(),
            "valor_comissao": to_float(get(row, "VLRCOMIS")),
            "base_comissao": to_float(get(row, "BASECOMIS")),
            "valor_bruto_item": to_float(get(row, "VBRUTOITEM")),
            "perc_total": to_float(get(row, "PERCTOTAL")),
            "nota_fiscal": str(get(row, "NOTAFISCAL") or ""),
            "serie_nf": str(get(row, "SERIENF") or ""),
            "item_nf": str(get(row, "ITEMNF") or ""),
            "emissao_nf": to_date(get(row, "EMISSAONF")),
            "dt_vencto": to_date(get(row, "DTVENCTO")),
            "dt_baixa": to_date(get(row, "DTBAIXA")),
            "referencia": ref_raw,
            "situacao": str(get(row, "SITUACAO") or ""),
            "recorrente": get(row, "RECORRENTE"),
            "contrato": str(get(row, "CONTRATO") or ""),
            "dt_contrato": to_date(get(row, "DTCONTRATO")),
            "tipo_prov": str(get(row, "TIPOPROV") or ""),
            "qtd_mes_cli": to_float(get(row, "QTDMESCLI")),
        })

    docs: List[Dict[str, Any]] = []
    for ref_key in sorted(by_ref_clientes.keys()):
        clientes = by_ref_clientes[ref_key]
        cnpjs_filial = [
            {
                "cod_empresa": v["cod_empresa"],
                "cnpj": v["cnpj"],
                "total": round(v["total"], 2),
                "cod_fornecedor": "",
                "num_pedido": "",
                "razao_social": v["razao_social"],
                "endereco": "",
            }
            for v in by_ref_filiais[ref_key].values()
        ]
        total_geral = round(sum((c.get("valor_comissao") or 0) for c in clientes), 2)
        docs.append({
            "tipo": "REAL",
            "competencia": ref_key,
            "total_geral": total_geral,
            "cnpjs_filial": cnpjs_filial,
            "clientes": clientes,
            "source_file": xlsx_path.name,
            "source_email_subject": source_meta.get("subject", ""),
            "source_email_from": source_meta.get("from", ""),
            "source_email_date": br_email_date_to_iso(source_meta.get("date", "")),
            "source_email_date_raw": source_meta.get("date", ""),
            "n_linhas": len(clientes),
        })
    return docs


def parse_xlsx(xlsx_path: Path, source_meta: Dict[str, str]) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    if "Plan2" not in wb.sheetnames or "Plan1" not in wb.sheetnames:
        raise ValueError(f"esperava Plan1+Plan2 em {xlsx_path.name}")

    # Plan2 — resumo por filial Totvs
    p2 = wb["Plan2"]
    cnpjs_filial: List[Dict[str, Any]] = []
    total_geral: Optional[float] = None
    for row in p2.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        first = str(row[0]).strip()
        if first.startswith("Soma de"):
            continue
        if first == "EMPRESA":
            continue
        if first == "Total Geral":
            try:
                total_geral = float(row[2]) if row[2] is not None else None
            except Exception:
                total_geral = None
            continue
        # linha de filial: row = (cod_empresa, cnpj, total, cod_forn, num_pedido, razao_social, endereco)
        try:
            if row[1] and row[2] is not None:
                cnpjs_filial.append({
                    "cod_empresa": str(row[0]) if row[0] is not None else "",
                    "cnpj": str(row[1]),
                    "total": round(float(row[2]), 2),
                    "cod_fornecedor": str(row[3]) if len(row) > 3 and row[3] is not None else "",
                    "num_pedido": str(row[4]) if len(row) > 4 and row[4] is not None else "",
                    "razao_social": str(row[5]) if len(row) > 5 and row[5] is not None else "",
                    "endereco": str(row[6]) if len(row) > 6 and row[6] is not None else "",
                })
        except Exception:
            pass

    # Plan1 — base detalhada por cliente final
    p1 = wb["Plan1"]
    headers = {p1.cell(1, c).value: c for c in range(1, p1.max_column + 1) if p1.cell(1, c).value}
    needed = ["EMPRESA", "CNPJ", "CODCLIENTE", "CLIENTE", "DESCPROD", "PRODUTO",
              "VLRCOMIS", "BASECOMIS", "VBRUTOITEM", "PERCTOTAL",
              "NOTAFISCAL", "SERIENF", "ITEMNF", "EMISSAONF", "DTVENCTO", "DTBAIXA",
              "REFERENCIA", "SITUACAO", "RECORRENTE", "CONTRATO", "DTCONTRATO",
              "TIPOPROV", "QTDMESCLI"]
    def get(row, key):
        c = headers.get(key)
        if c is None:
            return None
        return p1.cell(row, c).value

    def to_date(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10] if v else None

    def to_float(v):
        if v is None:
            return None
        try:
            return round(float(v), 2)
        except Exception:
            return None

    clientes: List[Dict[str, Any]] = []
    referencias = set()
    for r in range(2, p1.max_row + 1):
        if get(r, "VLRCOMIS") is None and get(r, "CLIENTE") is None:
            continue
        ref = to_date(get(r, "REFERENCIA"))
        if ref:
            referencias.add(ref)
        clientes.append({
            "cnpj_filial": str(get(r, "CNPJ") or ""),
            "cod_cliente": str(get(r, "CODCLIENTE") or ""),
            "cliente": str(get(r, "CLIENTE") or "").strip(),
            "produto_cod": str(get(r, "PRODUTO") or ""),
            "produto_desc": str(get(r, "DESCPROD") or "").strip(),
            "valor_comissao": to_float(get(r, "VLRCOMIS")),
            "base_comissao": to_float(get(r, "BASECOMIS")),
            "valor_bruto_item": to_float(get(r, "VBRUTOITEM")),
            "perc_total": to_float(get(r, "PERCTOTAL")),
            "nota_fiscal": str(get(r, "NOTAFISCAL") or ""),
            "serie_nf": str(get(r, "SERIENF") or ""),
            "item_nf": str(get(r, "ITEMNF") or ""),
            "emissao_nf": to_date(get(r, "EMISSAONF")),
            "dt_vencto": to_date(get(r, "DTVENCTO")),
            "dt_baixa": to_date(get(r, "DTBAIXA")),
            "referencia": ref,
            "situacao": str(get(r, "SITUACAO") or ""),
            "recorrente": get(r, "RECORRENTE"),
            "contrato": str(get(r, "CONTRATO") or ""),
            "dt_contrato": to_date(get(r, "DTCONTRATO")),
            "tipo_prov": str(get(r, "TIPOPROV") or ""),
            "qtd_mes_cli": to_float(get(r, "QTDMESCLI")),
        })

    # competencia: usa REFERENCIA da maioria das linhas SE for data válida YYYY-MM-DD;
    # fallback para nome do arquivo (cobre PROVISAO, onde REFERENCIA vem como string)
    competencia = None
    if referencias:
        from collections import Counter
        valid = [c for c in referencias if c and re.fullmatch(r"\d{4}-\d{2}-\d{2}", c)]
        if valid:
            competencia = Counter(valid).most_common(1)[0][0]
    if not competencia:
        competencia = parse_competencia_from_filename(xlsx_path.name)

    tipo = detect_tipo(source_meta.get("subject", ""), xlsx_path.name)

    return {
        "tipo": tipo,
        "competencia": competencia,
        "total_geral": round(total_geral, 2) if total_geral is not None else None,
        "cnpjs_filial": cnpjs_filial,
        "clientes": clientes,
        "source_file": xlsx_path.name,
        "source_email_subject": source_meta.get("subject", ""),
        "source_email_from": source_meta.get("from", ""),
        "source_email_date": br_email_date_to_iso(source_meta.get("date", "")),
        "source_email_date_raw": source_meta.get("date", ""),
        "n_linhas": len(clientes),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="indir", type=Path, default=DEFAULT_IN,
                    help="diretório com .eml e/ou .xlsx (default: ~/Documents/Easy_Financas/totvs/raw)")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT,
                    help="caminho do snapshot.json")
    ap.add_argument("--drive-manifest", dest="drive_manifest", type=Path, default=None,
                    help="JSON com lista de filenames atualmente em Drive. "
                         "Quando presente, ativa modo 'Drive=fonte da verdade': "
                         "snapshot é regenerado do zero contendo apenas docs cujo nome "
                         "(do .eml ou .xlsx) esteja na manifest. "
                         "Quando ausente, modo legado: snapshot é cumulativo sobre tudo em --in.")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    indir: Path = args.indir
    out: Path = args.out
    if not indir.exists():
        log(f"[warn] diretório {indir} não existe; criando vazio", args.quiet)
        indir.mkdir(parents=True, exist_ok=True)
        # ainda assim grava snapshot vazio
    out.parent.mkdir(parents=True, exist_ok=True)

    # Carrega manifest do Drive (se fornecido). Em modo Drive=fonte-da-verdade,
    # snapshot é regenerado do zero a cada rodada — só inclui docs cujo nome de
    # origem (.eml ou .xlsx) consta no manifest.
    drive_filenames: Optional[set] = None
    drive_manifest_meta: Dict[str, Any] = {}
    if args.drive_manifest is not None:
        if args.drive_manifest.exists():
            try:
                m = json.loads(args.drive_manifest.read_text(encoding="utf-8"))
                files = m.get("files", [])
                drive_filenames = set(files)
                drive_manifest_meta = {
                    "path": str(args.drive_manifest),
                    "generated_at": m.get("generated_at"),
                    "drive_folder_id": m.get("drive_folder_id"),
                    "n_files": len(drive_filenames),
                }
                log(f"[drive] manifest carregado: {len(drive_filenames)} arquivo(s) em Drive", args.quiet)
            except Exception as e:
                log(f"[warn] manifest ilegível ({e}); usando modo legado (cumulativo)", args.quiet)
                drive_filenames = None
        else:
            log(f"[warn] --drive-manifest passado mas {args.drive_manifest} não existe; modo legado", args.quiet)

    # carrega snapshot existente. Em modo Drive=fonte-da-verdade, regenera do zero
    # (snapshot reflete exatamente o estado do Drive). Em modo legado, é cumulativo.
    existing: Dict[str, Any] = {"documentos": [], "_meta": {}}
    if drive_filenames is None and out.exists():
        try:
            existing = json.loads(out.read_text(encoding="utf-8"))
            if "documentos" not in existing:
                existing["documentos"] = []
        except Exception as e:
            log(f"[warn] snapshot existente ilegível ({e}); recomeçando", args.quiet)
            existing = {"documentos": [], "_meta": {}}

    seen_hashes = {d.get("hash") for d in existing["documentos"] if d.get("hash")}

    # itera arquivos: .eml primeiro (extrai .xlsx), depois .xlsx soltos
    work_dir = indir / ".extracted"
    work_dir.mkdir(parents=True, exist_ok=True)

    def in_drive(name: str) -> bool:
        """Retorna True se o filename está em Drive (ou se não há manifest = modo legado)."""
        if drive_filenames is None:
            return True
        return name in drive_filenames

    # Extrai TODOS os .eml em raw/ (cheap+idempotente graças ao outp.exists()).
    # Depois filtra os pairs pelo manifest do Drive: match é por nome do anexo
    # .xlsx original (não pelo nome do .eml), porque é o .xlsx que vai pro Drive.
    all_pairs: List[Tuple[Path, Dict[str, str]]] = []
    extracted_this_run: set[Path] = set()  # caminhos canônicos esperados após esta rodada
    for eml in sorted(indir.glob("*.eml")):
        try:
            for xlsx, meta in extract_xlsx_from_eml(eml, work_dir):
                all_pairs.append((xlsx, meta))
                extracted_this_run.add(xlsx.resolve())
        except Exception as e:
            log(f"[warn] erro extraindo {eml.name}: {e}", args.quiet)
    for xlsx in sorted(indir.glob("*.xlsx")):
        # arquivos xlsx soltos (sem .eml ao redor)
        all_pairs.append((xlsx, {"subject": xlsx.name, "from": "", "date": "",
                                 "original_xlsx_filename": xlsx.name}))

    # Aplica filtro do manifest: pair sobrevive se seu nome canônico (anexo do
    # .eml ou nome do .xlsx solto) consta no manifest.
    pairs: List[Tuple[Path, Dict[str, str]]] = []
    drive_skipped = 0
    for xlsx, meta in all_pairs:
        canonical = meta.get("original_xlsx_filename") or xlsx.name
        if in_drive(canonical):
            pairs.append((xlsx, meta))
        else:
            drive_skipped += 1
    if drive_filenames is not None and drive_skipped > 0:
        log(f"[drive] {drive_skipped} doc(s) locais ignorados (não estão em Drive)", args.quiet)

    added = 0
    skipped = 0
    errors = 0
    for xlsx, meta in pairs:
        try:
            h = sha256_file(xlsx)
            if h in seen_hashes:
                skipped += 1
                continue
            # Dispatch: FATURAMENTO (multi-mês, tipo=REAL) vs comissão padrão (1 doc)
            if is_faturamento_file(xlsx.name):
                new_docs = parse_faturamento_xlsx(xlsx, meta)
            else:
                new_docs = [parse_xlsx(xlsx, meta)]
            for doc in new_docs:
                doc["hash"] = h
                existing["documentos"].append(doc)
            seen_hashes.add(h)
            added += len(new_docs)
            if len(new_docs) == 1:
                doc = new_docs[0]
                log(f"[add] {doc['tipo']:12} comp={doc['competencia']} R$ {doc['total_geral']:>12,.2f} | {xlsx.name}", args.quiet)
            else:
                total_real = sum(d.get("total_geral", 0) or 0 for d in new_docs)
                log(f"[add] {len(new_docs)} docs REAL ({new_docs[0]['competencia']}..{new_docs[-1]['competencia']}) R$ {total_real:>12,.2f} | {xlsx.name}", args.quiet)
        except Exception as e:
            errors += 1
            log(f"[err] {xlsx.name}: {e}", args.quiet)

    # Dedup REAL vs BASE/COMPLEMENTAR: para meses cobertos pelos relatórios
    # mensais (BASE ou COMPLEMENTAR), o REAL agregado do FATURAMENTO é
    # redundante (e seria triple-counting). Mantém REAL só nos meses históricos
    # sem cobertura mensal.
    covered_months = {d["competencia"] for d in existing["documentos"]
                      if d.get("tipo") in ("BASE", "COMPLEMENTAR", "PROVISAO")
                      and d.get("competencia")}
    before_n = len(existing["documentos"])
    existing["documentos"] = [
        d for d in existing["documentos"]
        if not (d.get("tipo") == "REAL" and d.get("competencia") in covered_months)
    ]
    real_dropped = before_n - len(existing["documentos"])
    if real_dropped > 0:
        log(f"[dedup] {real_dropped} doc(s) REAL removidos (já cobertos por BASE/COMPLEMENTAR/PROVISAO)", args.quiet)

    # GC do .extracted/: remove arquivos que NÃO foram (re)extraídos nesta
    # rodada. Só remove se o hash do conteúdo já está no snapshot (segurança:
    # nunca deletar um documento que ainda não foi processado). Arquivos
    # vivos = os que extract_xlsx_from_eml produziu (nomes canônicos atuais).
    # Estratégia: roda só faz sentido quando havia .eml(s) na pasta (senão
    # extracted_this_run estaria vazio e zeraria tudo erroneamente).
    gc_removed = 0
    gc_skipped_perms = 0
    gc_orphans = 0
    if extracted_this_run:
        for x in sorted(work_dir.glob("*.xlsx")):
            xr = x.resolve()
            if xr in extracted_this_run:
                continue  # arquivo da rodada atual, mantém
            try:
                h = sha256_file(x)
            except Exception:
                continue
            if h not in seen_hashes:
                gc_orphans += 1
                continue  # conteúdo não está no snapshot — mantém para revisão
            try:
                x.unlink()
                gc_removed += 1
            except PermissionError:
                gc_skipped_perms += 1
            except Exception as e:
                log(f"[gc warn] não removi {x.name}: {e}", args.quiet)
    if gc_removed > 0:
        log(f"[gc] removidas {gc_removed} duplicatas órfãs em .extracted/", args.quiet)
    if gc_skipped_perms > 0:
        log(f"[gc] {gc_skipped_perms} arquivo(s) bloqueados pelo iCloud — rode no Mac nativo p/ limpar", args.quiet)
    if gc_orphans > 0:
        log(f"[gc] {gc_orphans} arquivo(s) sem doc no snapshot (mantidos para revisão)", args.quiet)

    existing["_meta"] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_dir": str(indir),
        "n_documentos": len(existing["documentos"]),
        "tool": "fetch-totvs.py v2",
        "mode": "drive_source_of_truth" if drive_filenames is not None else "legacy_cumulative",
        "drive_manifest": drive_manifest_meta if drive_manifest_meta else None,
    }

    # write atomic: tmp + rename
    tmp = out.with_suffix(out.suffix + ".tmp")
    tmp.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp, out)
    log(f"[done] {added} novos | {skipped} já no snapshot | {errors} erros | total documentos: {len(existing['documentos'])}", args.quiet)
    log(f"[done] snapshot: {out}", args.quiet)


if __name__ == "__main__":
    main()
